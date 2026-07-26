#!/usr/bin/env python3
"""Build the official-format PVRIG Final50 submission workbook from DOCX."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import xml.etree.ElementTree as ET
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


WORD_NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
}

EXPECTED_HEADERS = [
    "序号",
    "抗体名称",
    "抗体形式",
    "设计类型",
    "VH/VHH\n氨基酸序列",
    "VL\n氨基酸序列",
    "起始分子名称/来源",
    "起始分子\nVH/VHH序列",
    "起始分子\nVL序列",
    "设计说明",
    "自检结果",
    "模型预测得分/推荐依据",
    "推荐排序(Rank)",
]

PARENT_MAP = {
    "positive_pose_source_case02_pos_01_PVRIG-151_HR151": (
        "PVRIG-151_HR151",
        6,
    ),
    "positive_pose_source_case02_pos_10_151H7": ("151H7", 98),
    "positive_pose_source_case02_pos_11_151H8": ("151H8", 99),
    "positive_pose_source_case02_pos_04_PVRIG-38": ("PVRIG-38", 4),
}

FIXED_POSE_PARENT_MAP = {
    "case02_pos_01_PVRIG-151_HR151": (
        "PVRIG-151/HR-151",
        "WO2021180205A1（SEQ ID NO:6）",
    ),
    "case02_pos_10_151H7": (
        "151H7",
        "WO2021180205A1（SEQ ID NO:98）",
    ),
    "case02_pos_04_PVRIG-38": (
        "PVRIG-38",
        "WO2021180205A1（SEQ ID NO:4）",
    ),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template-docx", required=True, type=Path)
    parser.add_argument("--joined-tsv", required=True, type=Path)
    parser.add_argument("--parent-fasta", required=True, type=Path)
    parser.add_argument("--fixed-pose-audit-tsv", required=True, type=Path)
    parser.add_argument("--freeze-receipt", required=True, type=Path)
    parser.add_argument("--output-xlsx", required=True, type=Path)
    parser.add_argument("--archive-xlsx", type=Path)
    parser.add_argument("--receipt-json", required=True, type=Path)
    return parser.parse_args()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def extract_template_headers(path: Path) -> list[str]:
    with ZipFile(path) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))
    table = root.find(".//w:tbl", WORD_NS)
    if table is None:
        raise ValueError("template DOCX contains no table")
    first_row = table.find("./w:tr", WORD_NS)
    if first_row is None:
        raise ValueError("template table contains no rows")
    headers: list[str] = []
    for cell in first_row.findall("./w:tc", WORD_NS):
        parts: list[str] = []
        for paragraph in cell.findall(".//w:p", WORD_NS):
            text = "".join(
                node.text or "" for node in paragraph.findall(".//w:t", WORD_NS)
            ).strip()
            if text:
                parts.append(text)
        headers.append("\n".join(parts))
    return headers


def parse_fasta(path: Path) -> dict[str, dict[str, str]]:
    records: dict[str, dict[str, str]] = {}
    header: str | None = None
    parts: list[str] = []
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line:
            continue
        if line.startswith(">"):
            if header is not None:
                _store_fasta(records, header, "".join(parts))
            header = line[1:]
            parts = []
        else:
            parts.append(line)
    if header is not None:
        _store_fasta(records, header, "".join(parts))
    return records


def _store_fasta(
    records: dict[str, dict[str, str]], header: str, sequence: str
) -> None:
    fields = header.split("|")
    name = fields[0]
    metadata = {"name": name, "sequence": sequence}
    for field in fields[1:]:
        if "=" in field:
            key, value = field.split("=", 1)
            metadata[key] = value
    records[name] = metadata


def as_float(value: str) -> float:
    return float(value)


def as_bool(value: str) -> bool:
    return value.strip().lower() in {"1", "true", "yes", "pass", "passed"}


def design_type(row: dict[str, str]) -> str:
    if row["parent_cluster"] in PARENT_MAP:
        return "优化改造"
    if row["parent_cluster"].startswith("GENERATED_"):
        return "从头设计"
    raise ValueError(f"unmapped design provenance: {row['parent_cluster']}")


def parent_fields(
    row: dict[str, str], parents: dict[str, dict[str, str]]
) -> tuple[str, str]:
    if design_type(row) == "从头设计":
        return "不适用", "不适用"
    name, seq_id = PARENT_MAP[row["parent_cluster"]]
    parent = parents[name]
    source = f"{name} / WO2021180205A1（SEQ ID NO:{seq_id}）"
    return source, parent["sequence"]


def design_description(row: dict[str, str]) -> str:
    if design_type(row) == "从头设计":
        if row["route"] == "rfantibody":
            return (
                "从头设计（新骨架/全新CDR区）；采用RFantibody/RFdiffusion"
                "在PVRIG–PVRL2功能界面约束下生成VHH骨架，并使用"
                "ProteinMPNN设计CDR1、CDR2和CDR3序列；随后完成双构象、"
                "4个随机种子对接及统一序列和可开发性筛选。"
            )
        if row["route"] == "fixed_pose_mpnn":
            parent_id = row["_fixed_pose_parent_id"]
            parent_name, parent_source = FIXED_POSE_PARENT_MAP[parent_id]
            return (
                f"从头设计（全新CDR区）；保留公开专利阳性VHH {parent_name}"
                f"〔{parent_source}〕的framework，并以其计算获得的PVRIG"
                "结合pose作为结构约束，使用ProteinMPNN对CDR1、CDR2和"
                "CDR3全部重新设计；未直接沿用任何已知阳性抗体的完整CDR；"
                "随后完成双构象、4个随机种子对接及统一序列和可开发性筛选。"
            )
        raise ValueError(f"unmapped de-novo design route: {row['route']}")
    name, _seq_id = PARENT_MAP[row["parent_cluster"]]
    return (
        f"优化改造；以公开专利分子{name}为可追溯起始scaffold/pose，"
        "主要改造CDR1、CDR2、CDR3及结合界面序列；三个对应CDR对"
        "全部阳性参照identity均低于80%，并经4个随机种子×2个构象"
        "对接、阻断几何及统一可开发性筛选。"
    )


def nglycosylation_annotation(row: dict[str, str]) -> str:
    sequence = row["sequence"]
    cdr_ranges: list[tuple[int, int, int]] = []
    for index in (1, 2, 3):
        cdr = row[f"cdr{index}"]
        start = sequence.find(cdr)
        assert start >= 0, f"{row['submission_id']} CDR{index} not in sequence"
        cdr_ranges.append((index, start, start + len(cdr)))

    annotations: list[str] = []
    for match in re.finditer(r"N[^P][ST]", sequence):
        start = match.start()
        end = match.end()
        location = "非CDR区"
        for index, cdr_start, cdr_end in cdr_ranges:
            if cdr_start <= start and end <= cdr_end:
                location = f"位于CDR{index}"
                break
            if start < cdr_end and end > cdr_start:
                location = f"跨CDR{index}/FR邻接区域"
                break
        annotations.append(f"{match.group()}@{start + 1}（{location}）")

    if not annotations:
        return "全序列N-X-S/T motif=无"
    if all("位于CDR" not in item for item in annotations):
        return (
            "CDR内未见完整N-X-S/T；但全序列存在"
            f"{'、'.join(annotations)}，列为潜在糖基化风险"
        )
    return (
        f"全序列存在{'、'.join(annotations)}，"
        "其中CDR内motif列为内部高风险"
    )


def self_check(row: dict[str, str]) -> str:
    identities = [
        as_float(row[f"positive_max_positive_cdr{i}_identity"]) for i in (1, 2, 3)
    ]
    team_identities = [
        as_float(row[f"team_max_team_cdr{i}_identity"]) for i in (1, 2, 3)
    ]
    assert row["official_validator_pass"] == "true"
    assert row["positive_all_positive_corresponding_cdrs_below_0p80"] == "true"
    assert max(identities) < 0.80
    cys_count = int(row["uniform_seq_cys_count"])
    invalid_aa = int(row["uniform_seq_invalid_aa_count"])
    fusion_hard_fail = as_bool(row["prefusion_fusion_hard_fail"])
    internal_hard_fail = bool(row["primary_hard_fail_reasons"])
    return (
        "序列合规自检PASS：官方ab-data-validator通过；IMGT/ANARCI完整；"
        f"长度{len(row['sequence'])} aa；非法氨基酸={invalid_aa}；"
        f"CDR1/CDR2/CDR3对阳性参照的最大identity="
        f"{identities[0] * 100:.1f}%/{identities[1] * 100:.1f}%/"
        f"{identities[2] * 100:.1f}%，均低于80%；"
        f"队内最近邻CDR1/CDR2/CDR3 identity="
        f"{team_identities[0] * 100:.1f}%/{team_identities[1] * 100:.1f}%/"
        f"{team_identities[2] * 100:.1f}%（用于组合多样性复核）；"
        f"Cys={cys_count}；{nglycosylation_annotation(row)}；"
        f"融合预检查硬冲突={'有' if fusion_hard_fail else '无'}；"
        f"内部开发性硬风险={'有，需复核' if internal_hard_fail else '无'}；"
        f"SHA256={row['sequence_sha256'][:16]}…"
    )


REVIEW_REASON_ZH = {
    "ELEVATED_SURFACE_HYDROPHOBIC_PATCH": "表面疏水斑块偏高",
    "ABNATIV_BELOW_0.70": "AbNatiV VHH分数低于0.70",
    "EXPOSED_ACID_CLIPPING_ROWS_GE_4": "暴露酸敏感位点较多",
    "BORDERLINE_SINGLE_DOMAIN": "单域适配性处于边界",
    "TNP_FLAGS_MISSING": "部分TNP指标技术性缺失",
    "UNUSUAL_PI": "等电点处于非常规范围",
    "PATENT_SUCCESS_SCAFFOLD_POOR_SINGLE_DOMAIN_REVIEW": (
        "公开成功scaffold但单域适配性需复核"
    ),
    "NON_CDR_N_GLYCOSYLATION_MOTIF": "非CDR区存在N-糖基化motif",
    "SAPIENS_BELOW_0.70": "Sapiens自相似分数低于0.70",
    "TNP_RED_FLAG_1": "1项TNP红色风险",
    "TNP_AMBER_FLAGS_1": "1项TNP黄色警示",
    "TNP_AMBER_FLAGS_2": "2项TNP黄色警示",
    "INSTABILITY_INDEX_GE_40": "instability index不低于40",
    "EXTREME_SURFACE_HYDROPHOBIC_PATCH": "表面疏水斑块过大",
    "CDR_N_GLYCOSYLATION_MOTIF": "CDR区存在N-糖基化motif",
}


def translate_risk_reasons(row: dict[str, str]) -> str:
    reasons = [
        item
        for field in ("primary_hard_fail_reasons", "primary_review_reasons")
        for item in row[field].split(";")
        if item
    ]
    if not reasons:
        return "未发现达到内部阈值的主要开发性警示"
    translated: list[str] = []
    for reason in reasons:
        if reason == "INSTABILITY_INDEX_GE_40":
            translated.append(
                f"instability index={as_float(row['uniform_seq_instability_index']):.3f}"
                "（内部警示阈值40）"
            )
        else:
            translated.append(REVIEW_REASON_ZH.get(reason, reason))
    return "、".join(translated)


def recommendation(row: dict[str, str]) -> str:
    grade_map = {
        "A_LOWER_RISK": "A（低风险）",
        "B_REVIEW": "B（需复核）",
        "C_HIGH_RISK": "C（高风险储备）",
    }
    assert row["source_ranked_blocker_class"] == "CONSENSUS_BLOCKER_LIKE_A"
    fusion_text = (
        "融合兼容性窄范围预检查未见硬性冲突"
        if not as_bool(row["prefusion_fusion_hard_fail"])
        else "融合兼容性预检查发现硬性冲突"
    )
    epitope = row["epitope_epitope_cluster_id"].replace("EPI_", "")
    c_tail = (
        "；该候选仅作为高几何分、高开发风险储备，不建议列入前10优先"
        if row["primary_developability_grade"] == "C_HIGH_RISK"
        else ""
    )
    return (
        "双构象多种子阻断几何为A类共识；"
        f"严格/宽松seed支持为{row['source_ranked_strict_seed_count']}/"
        f"{row['source_ranked_broad_seed_count']}；"
        f"pose稳健性为{as_float(row['source_ranked_pose_robustness_score']):.2f}；"
        f"阻断共识为{as_float(row['source_ranked_blocking_consensus_score']):.2f}；"
        f"双构象一致性为"
        f"{as_float(row['source_ranked_dual_reference_agreement_fraction']) * 100:.1f}%；"
        f"结合位点接触模式为第{int(epitope)}类；"
        f"开发性内部等级为{grade_map[row['primary_developability_grade']]}，"
        f"依据：{translate_risk_reasons(row)}；{fusion_text}{c_tail}。"
        "以上0–100分均为内部归一化计算代理分，不代表实验阻断概率、"
        "BLI响应、表达量、纯度、Kd或IC50。"
    )


def build_workbook(headers: list[str], rows: list[dict[str, str]], parents: dict[str, dict[str, str]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "抗体提交表"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:M{len(rows) + 1}"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border
    sheet.row_dimensions[1].height = 46

    design_counts: Counter[str] = Counter()
    for excel_row, row in enumerate(rows, 2):
        dtype = design_type(row)
        design_counts[dtype] += 1
        parent_name, parent_sequence = parent_fields(row, parents)
        values = [
            int(row["competition_rank_1_50"]),
            row["submission_id"],
            "VHH纳米抗体",
            dtype,
            row["sequence"],
            "不适用",
            parent_name,
            parent_sequence,
            "不适用",
            design_description(row),
            self_check(row),
            recommendation(row),
            int(row["competition_rank_1_50"]),
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(excel_row, column, value)
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if column in {1, 2, 3, 4, 6, 9, 13} else "left",
                wrap_text=True,
            )
            if column in {5, 8}:
                cell.number_format = "@"
                cell.font = Font(name="Consolas", size=9)
            else:
                cell.font = Font(name="Microsoft YaHei", size=9)
        sheet.row_dimensions[excel_row].height = 155

    widths = {
        "A": 7,
        "B": 19,
        "C": 15,
        "D": 12,
        "E": 68,
        "F": 12,
        "G": 34,
        "H": 68,
        "I": 12,
        "J": 52,
        "K": 58,
        "L": 58,
        "M": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    antibody_validation = DataValidation(
        type="list", formula1='"IgG单克隆抗体,VHH纳米抗体"', allow_blank=False
    )
    design_validation = DataValidation(
        type="list", formula1='"从头设计,优化改造"', allow_blank=False
    )
    sheet.add_data_validation(antibody_validation)
    sheet.add_data_validation(design_validation)
    antibody_validation.add(f"C2:C{len(rows) + 1}")
    design_validation.add(f"D2:D{len(rows) + 1}")

    yellow = PatternFill("solid", fgColor="FFF2CC")
    red = PatternFill("solid", fgColor="F4CCCC")
    sheet.conditional_formatting.add(
        f"L2:L{len(rows)+1}",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("开发性内部等级为B",L2))'], fill=yellow
        ),
    )
    sheet.conditional_formatting.add(
        f"L2:L{len(rows)+1}",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("开发性内部等级为C",L2))'], fill=red
        ),
    )

    sheet.print_title_rows = "1:1"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = True
    sheet.oddFooter.center.text = "PVRIG Final50 — official template fields"
    sheet.oddFooter.right.text = "Page &P / &N"

    assert design_counts == Counter({"从头设计": 27, "优化改造": 23})
    return workbook


def verify_workbook(path: Path, rows: list[dict[str, str]], headers: list[str]) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet = workbook["抗体提交表"]
    actual_headers = [sheet.cell(1, column).value for column in range(1, 14)]
    assert actual_headers == headers
    assert sheet.max_row == 51
    assert sheet.max_column == 13
    for index, source in enumerate(rows, 2):
        assert sheet.cell(index, 1).value == int(source["competition_rank_1_50"])
        assert sheet.cell(index, 2).value == source["submission_id"]
        assert sheet.cell(index, 5).value == source["sequence"]
        assert sheet.cell(index, 13).value == int(source["competition_rank_1_50"])
        self_check_text = sheet.cell(index, 11).value
        recommendation_text = sheet.cell(index, 12).value
        assert "CDR1/CDR2/CDR3对阳性参照的最大identity=" in self_check_text
        assert "%" in self_check_text
        assert "机制rank" not in recommendation_text.lower()
        assert "CONSENSUS_BLOCKER_LIKE_A" not in recommendation_text
        assert "contact cluster" not in recommendation_text.lower()
        assert "以上0–100分均为内部归一化计算代理分" in recommendation_text
        if source["route"] == "fixed_pose_mpnn":
            description = sheet.cell(index, 10).value
            assert "公开专利阳性VHH" in description
            assert "framework" in description
            assert "计算获得的PVRIG结合pose" in description
            assert "未直接沿用任何已知阳性抗体的完整CDR" in description
        rank = int(source["competition_rank_1_50"])
        if rank in {26, 32, 34}:
            assert "NVT@58" in self_check_text
            assert "跨CDR2/FR邻接区域" in self_check_text
        if rank == 50:
            assert "NLS@101" in self_check_text
            assert "位于CDR3" in self_check_text


def main() -> None:
    args = parse_args()
    template_headers = extract_template_headers(args.template_docx)
    normalized = ["".join(header.split()) for header in template_headers]
    expected_normalized = ["".join(header.split()) for header in EXPECTED_HEADERS]
    assert normalized == expected_normalized, (
        f"template headers changed: {template_headers}"
    )
    headers = EXPECTED_HEADERS
    with args.joined_tsv.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle, delimiter="\t"))
    rows.sort(key=lambda row: int(row["competition_rank_1_50"]))
    assert len(rows) == 50
    assert [int(row["competition_rank_1_50"]) for row in rows] == list(range(1, 51))
    assert len({row["sequence"] for row in rows}) == 50
    with args.fixed_pose_audit_tsv.open(newline="", encoding="utf-8-sig") as handle:
        fixed_pose_audit_rows = list(csv.DictReader(handle, delimiter="\t"))
    fixed_pose_audit = {
        row["candidate_id"]: row for row in fixed_pose_audit_rows
    }
    assert len(fixed_pose_audit) == len(fixed_pose_audit_rows)
    expected_fixed_pose_ids = {
        row["candidate_id"] for row in rows if row["route"] == "fixed_pose_mpnn"
    }
    assert len(expected_fixed_pose_ids) == 15
    assert set(fixed_pose_audit) == expected_fixed_pose_ids
    for candidate_id in sorted(expected_fixed_pose_ids):
        audit = fixed_pose_audit[candidate_id]
        assert audit["audit_verdict"] == "PASS_ALL_THREE_CDRS_REDESIGNED"
        assert all(
            as_bool(audit[f"cdr{index}_fully_redesigned"]) for index in (1, 2, 3)
        )
        assert audit["evidence_path"]
        assert audit["evidence_sha256"]
        assert audit["parent_id"] in FIXED_POSE_PARENT_MAP
        matching = [row for row in rows if row["candidate_id"] == candidate_id]
        assert len(matching) == 1
        matching[0]["_fixed_pose_parent_id"] = audit["parent_id"]
        matching[0]["_fixed_pose_pose_id"] = audit["pose_id"]
    parents = parse_fasta(args.parent_fasta)
    for name, _seq_id in PARENT_MAP.values():
        assert name in parents

    freeze_receipt = json.loads(args.freeze_receipt.read_text(encoding="utf-8"))
    assert freeze_receipt["state"] == "FINAL_FREEZE_COMPLETE"
    assert freeze_receipt["official_validator"]["passed"] == 50
    assert freeze_receipt["official_validator"]["failed"] == 0
    assert (
        freeze_receipt["p1"]["fixed_pose_all_three_cdrs_redesigned"]
        == "PASS_15_OF_15"
    )

    workbook = build_workbook(headers, rows, parents)
    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output_xlsx)
    verify_workbook(args.output_xlsx, rows, headers)

    if args.archive_xlsx:
        args.archive_xlsx.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(args.output_xlsx, args.archive_xlsx)
        assert sha256(args.output_xlsx) == sha256(args.archive_xlsx)

    receipt = {
        "schema_version": "pvrig_final50_official_submission_xlsx_v1_2",
        "state": "COMPLETE",
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "records": 50,
        "columns": 13,
        "design_type_counts": {"从头设计": 27, "优化改造": 23},
        "antibody_form": "VHH纳米抗体",
        "ranking": "frozen competition_rank_1_50; literal PVRIG_CAND_001-050",
        "template_headers_exact": True,
        "official_validator_passed": 50,
        "inputs": {
            str(args.template_docx): sha256(args.template_docx),
            str(args.joined_tsv): sha256(args.joined_tsv),
            str(args.parent_fasta): sha256(args.parent_fasta),
            str(args.fixed_pose_audit_tsv): sha256(args.fixed_pose_audit_tsv),
            str(args.freeze_receipt): sha256(args.freeze_receipt),
        },
        "outputs": {
            str(args.output_xlsx): sha256(args.output_xlsx),
            **(
                {str(args.archive_xlsx): sha256(args.archive_xlsx)}
                if args.archive_xlsx
                else {}
            ),
        },
        "assertions": {
            "workbook_rows_exact_50": True,
            "workbook_columns_exact_13": True,
            "rank_exact_1_to_50": True,
            "workbook_sequences_exact_joined_evidence": True,
            "all_positive_corresponding_cdrs_below_0p80": True,
            "official_validator_50_of_50": True,
            "mechanism_rank_absent_from_official_recommendation": True,
            "scores_explicitly_labeled_nonexperimental_proxies": True,
            "positive_cdr_identity_rendered_as_percent": True,
            "fixed_pose_all_three_cdrs_redesigned_15_of_15": True,
            "fixed_pose_positive_framework_pose_lineage_disclosed_15_of_15": True,
            "full_sequence_nglycosylation_motifs_reported": True,
        },
        "claim_boundary": (
            "Computational submission workbook; model scores are recommendation "
            "evidence, not measured expression, purity, BLI, Kd, IC50 or blocking."
        ),
    }
    args.receipt_json.parent.mkdir(parents=True, exist_ok=True)
    args.receipt_json.write_text(
        json.dumps(receipt, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(receipt, ensure_ascii=False, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
